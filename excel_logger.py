# excel_logger.py
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

USB_MOUNT = "/media/ruki/KIOXIA"
EXCEL_NAME = "attendance_log.xlsx"
LOCAL_PATH = "/home/ruki/offline_log.xlsx"

def _is_usb_available() -> bool:
    return os.path.ismount(USB_MOUNT)

def _target_path() -> str:
    return os.path.join(USB_MOUNT, EXCEL_NAME) if _is_usb_available() else LOCAL_PATH

def save_attendance_excel(name: str | None, card_id: str, dt: datetime | None = None):
    if dt is None:
        dt = datetime.now()

    date_str = dt.strftime("%Y-%m-%d")
    time_str = dt.strftime("%H:%M")
    excel_path = _target_path()

    # パス先のフォルダがなければ作成
    folder = os.path.dirname(excel_path)
    os.makedirs(folder, exist_ok=True)

    if not name:
        name = ""

    # ファイルがなければ新規作成
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "card_id", "date", "time1"])
    else:
        wb = load_workbook(excel_path)
        ws = wb.active

    # 同日に同カードIDの行があれば追記、なければ新規行
    found_row = None
    for row in ws.iter_rows(min_row=2):
        if row[1].value == card_id and row[2].value == date_str:
            found_row = row[0].row
            break

    if found_row:
        next_col = len([c for c in ws[found_row] if c.value is not None]) + 1
        ws[f"{get_column_letter(next_col)}{found_row}"] = time_str
        if not ws[f"A{found_row}"].value:
            ws[f"A{found_row}"] = name
    else:
        ws.append([name, card_id, date_str, time_str])

    wb.save(excel_path)
    print(f"[INFO] Excel 保存完了 → {excel_path}")
